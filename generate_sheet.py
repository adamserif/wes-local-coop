import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.writer.excel import save_virtual_workbook
from tempfile import NamedTemporaryFile

def generate_totals_sheet(shares, resp_df):
    '''
    Given a list of share prices and a cleaned up
    survey response dataframe,
    saves totals for each group as an Excel spreadsheet
    to be delivered by the Flask app
    '''

    coops = ["Produce", "Meat", "Eggs", "Bread", "Dairy", "Cheese", "Preserves", "Coffee" , "Tofu", "Seitan", "Mushroom"]

    raw = Workbook()
    raw_s = raw.active

    for r in dataframe_to_rows(resp_df , index=True, header=False):
        raw_s.append(r)

    raw.save('test.xlsx')
    alph = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    wb = Workbook()
    ws = wb.active
    ws.title = "Shares by group"

    totals = wb.create_sheet()
    totals.title = "Individual totals"

    contacts = wb.create_sheet()
    contacts.title = "Contacts"

    members = wb.create_sheet()
    members.title = "Members of each Co-Op"

    ws['A1'] = "Group count"
    #NAMES OF GROUP MEMBERS
    for i in range(2,8):
        ws.cell(row = 1, column = i).value = "Name "+str(i-1)

    #PRODUCE
    ws['H1'] = "Produce Shares"
    ws['I1'] = "Produce Charge"
    for i in range(1,7):
        ws.cell(row=1,column=9+i).value = "Split Percent" + str(i)

    #MEAT
    ws['P1'] = "Meat Shares"
    ws['Q1'] = "Meat Charge"
    for i in range(1,7):
        ws.cell(row=1,column=17+i).value = "Split Percent" + str(i)

    #EGGS
    ws['X1'] = "Egg Shares"
    ws['Y1'] = "Egg Charge"
    for i in range(1,7):
        ws.cell(row=1,column=25+i).value = "Split Percent" + str(i)

    #BREAD
    ws['AF1'] = "Bread Shares"
    ws['AG1'] = "Bread Charge"
    for i in range(1,7):
        ws.cell(row=1,column=33+i).value = "Split Percent" + str(i)

    #DAIRY
    ws['AN1'] = "Dairy Shares"
    ws['AO1'] = "Dairy Charge"
    for i in range(1,7):
        ws.cell(row=1,column=41+i).value = "Split Percent" + str(i)

    #CHEESE
    ws['AV1'] = "Cheese Shares"
    ws['AW1'] = "Cheese Charge"
    for i in range(1,7):
        ws.cell(row=1,column=49+i).value = "Split Percent" + str(i)

    #PRESERVES
    ws['BD1'] = "Preserves Shares"
    ws['BE1'] = "Preserves Charge"
    for i in range(1,7):
        ws.cell(row=1,column=57+i).value = "Split Percent" + str(i)

    #COFFEE
    ws['BL1'] = "Coffee Shares"
    ws['BM1'] = "Coffee Charge"
    for i in range(1,7):
        ws.cell(row=1,column=65+i).value = "Split Percent" + str(i)

    #TOFU
    ws['BT1'] = "Tofu Shares"
    ws['BU1'] = "Tofu Charge"
    for i in range(1,7):
        ws.cell(row=1,column=73+i).value = "Split Percent" + str(i)

    #SEITAN
    ws['CB1'] = "Seitan Shares"
    ws['CC1'] = "Seitan Charge"
    for i in range(1,7):
        ws.cell(row=1,column=81+i).value = "Split Percent" + str(i)

    #MUSHROOM
    ws['CJ1'] = "Mushroom Shares"
    ws['CK1'] = "Mushroom Charge"
    for i in range(1,7):
        ws.cell(row=1,column=89+i).value = "Split Percent" + str(i)

    #NOTES
    ws['CR1'] = "Notes"

    #TOTAL CHARGES
    ws['CS1'] = "Total Group Charge"
    for i in range(1,7):
        ws.cell(row=1,column=97+i).value = "Person " + str(i) + " charge"


    def find(x,l):
        for i in range(len(l)):
            if l[i].value == x:
                return i
        return -1


    #copy groups from raw data to new spreadsheet
    total_people = 0
    total_groups = 0
    e = True
    i = 2
    while e: #find total number of groups
        if raw_s.cell(row = i, column = 1).value == None:
            e = False
        else:
            total_groups += 1
            i += 1


    print("total number of groups: " + str(total_groups))


    #SHARES BY GROUP TAB
    for i in range(1,total_groups + 1):
        #initialize row in new workbook:
        for j in range(103):
            ws.cell(row = i+1 , column = j+1)
        new_row=list(ws.rows)[i][:]
        cur_row=list(raw_s.rows)[i][:]
        new_row[0].value = i
        #initialize group size
        notempty = True
        groupsize = 1
        k = 2
        #find group size
        while notempty:
            c = cur_row[k]
            k += 1
            if c.value == "" or c.value == None or c.value == float('nan'):
                notempty = False
            elif type(c.value) != str:
                notempty = False
            elif groupsize == 6:
                notempty = False
            else:
                groupsize += 1

        total_people += groupsize


        #copy group names:
        for p in range(groupsize):
            new_row[p+1].value = cur_row[p+1].value


        #apply individual charges for each coop
        raw_i = 19 #column T
        new_i = 7
        for coop in range(11):
            share = shares[coop] #get share price
            if cur_row[raw_i].value == None or cur_row[raw_i].value == 0: #pass if not participating
                pass
            else:
                numshares=float(cur_row[raw_i].value)
                new_row[new_i].value = numshares
                new_row[new_i+1].value = numshares*share #total charge
            raw_i += 14
            new_i += 8


    print("total number of people in co-op: " + str(total_people))
    #initialize cells in totals sheet
    for r in range(total_people + 1):
        for c in range(12):
            totals.cell(row = r+1, column = c+1)

    totals.cell(row = 1, column = 1).value = "Name"
    totals.cell(row = 1, column = 2).value = "Total"
    for coop in range(len(coops)):
        totals.cell(row = 1, column = 3 + coop).value = coops[coop]


    #copy totals and splitting percentages
    cnt = 1
    for i in range(1,total_groups+1):
        print("copying totals for group " + str(i))
        group=list(raw_s.rows)[i][1:8]
        charges = list(ws.rows)[i][:]
        notempty = True
        groupsize = 1
        k = 1
        while notempty:
            c = group[k]
            if c.value == None or c.value == " ":
                notempty = False
            elif type(c.value) != str:
                notempty = False
            elif groupsize == 6:
                notempty = False
            else:
                groupsize += 1
                k += 1
        group = list(raw_s.rows)[i][1:groupsize+1]
        group_cost = 0
        for person in range(1,groupsize+1):
            total_cost=0
            totals.cell(row = cnt + person, column = 1).value = group[person-1].value        #name
            for coop in range(11):
                #print coops[coop]
                #participating = raw_s.cell(row = i+1, column = 21+14*coop+person-1).value
                participating = list(raw_s.rows)[i][20+14*coop:26+14*coop]
                evensplit = raw_s.cell(row = i+1, column = 27+14*coop).value  # Yes = even split, No = uneven split
                nshares = float(raw_s.cell(row = i+1, column = 20+14*coop).value)      #number of shares
                if evensplit == "No":
                    #UNEVEN SPLITTING
                    percent = float(raw_s.cell(row = i+1, column = 28+14*coop+person-1).value)
                    cost = (0 if nshares is None else nshares)*(1 if shares[coop] is None else shares[coop]) * (0 if percent is None else percent)/100
                    total_cost+=cost
                elif nshares == 0.0 or participating[person-1].value == None or nshares == None: #not participating in this coop
                    cost = 0
                    percent = 0
                else:
                    #EVEN SPLITTING
                    n = 0
                    for p in participating:
                        if p.value == None:
                            pass
                        else:
                            n+=1
                    percent = 100.0/n
                    cost = nshares*shares[coop]/n
                    total_cost+=cost
                ws.cell(row = i+1,column = 10+8*coop+person-1).value = percent
                totals.cell(row = cnt+person, column = 3+coop).value = cost
            ws.cell(row = i + 1, column = 97+person).value = total_cost
            group_cost+=total_cost
        ws.cell(row = i + 1, column = 97).value = group_cost
        cnt += groupsize

    # Contact Info
    for j in range(1,7): # Header
        contacts.cell(row = 1, column = j).value = "Name" + str(j)
        contacts.cell(row = 1, column = j + 6).value = "Email" + str(j)
        contacts.cell(row = 1, column = j + 12).value = "WesID" + str(j)

    for i in range(1,total_groups+1):
        for j in range(1,19):
            contacts.cell(row = i + 1 , column = j).value = raw_s.cell(row = i + 1, column = j+1).value



    # Members of each co-op
    for coop in range(11):
        members.cell(row=1,column=coop+1).value = coops[coop]
        cnt = 0
        for i in range(total_people):
            paid = totals.cell(row = i+2, column = coop + 3).value
            if paid != 0:
                members.cell(row=cnt+2,column=coop+1).value = totals.cell(row=i+2,column=1).value
                cnt+=1

    emails = wb.create_sheet()
    emails.title = "Emails"

    # Create email lists for each co-op
    for coop in range(11):
        emails.cell(row=1,column=coop+1).value=coops[coop]
        cnt=2
        for i in range(2,total_groups+2):
            for person in range(6):
                name = ws.cell(row=i,column=2+person).value
                if name == None or name == " ": #name
                    pass
                else:
                    has = ws.cell(row=i,column=10+8*coop+person).value #splitting percentage
                    if has == None or has == 0:
                        pass
                    else:
                        emails.cell(row=cnt,column=1+coop).value = contacts.cell(row=i,column=7+person).value
                        cnt+=1

    wb.template = True
    wb.save("coop_totals.xlsx")